from openpyxl import load_workbook

# Replace 'your_file.xlsx' with the path to your Excel file
wb = load_workbook(filename='your_file.xlsx', data_only=False)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if cell.data_type == 'f':  # This is a formula
                try:
                    # Attempt to evaluate the formula
                    temp_val = ws.formula_attributes[cell.coordinate]
                except Exception as e:
                    print(f"Error in Sheet: {sheet_name}, Cell: {cell.coordinate}, Formula: {cell.value}, Error: {e}")
